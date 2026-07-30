import openpyxl, os

dest = r'd:\文旅局数据\文旅局'

files_to_check = [
    ('主模板', '咸丰县智慧旅游应用数据收集清单-文旅局上报数据模板(13).xlsx'),
    ('好声音', '文旅局上报数据-咸丰县好声音娱乐会所.xlsx'),
    ('嘉乐迪', r'嘉乐迪\咸丰县智慧旅游应用数据收集清单-文旅局上报数据模板-2(1).xlsx'),
    ('火塘民谣', r'火塘民谣 三人行电竞\咸丰县智慧旅游应用数据收集清单-文旅局上报数据-火塘民谣三人行电竞.xlsx'),
]

for label, rel_path in files_to_check:
    fp = os.path.join(dest, rel_path)
    if not os.path.exists(fp):
        continue
    
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    
    for sname in wb.sheetnames:
        if '娱乐' not in sname:
            continue
        ws = wb[sname]
        if ws.max_row is None:
            continue
        
        header_row = None
        headers = []
        for r in range(1, min(15, ws.max_row + 1)):
            vals = [str(c.value)[:20] if c.value else '' for c in ws[r]]
            if '序号' in vals:
                header_row = r
                headers = [str(c.value) if c.value else '' for c in ws[r]]
                break
        if header_row is None:
            continue
        
        data_start = header_row + 1
        first_vals = [str(c.value)[:20] if c.value else '' for c in ws[data_start]]
        if any('样例' in v for v in first_vals):
            data_start = header_row + 2
        
        print(f'\n===== {label} [{sname}] =====')
        for r in range(data_start, (ws.max_row or 50) + 1):
            vals = [c.value for c in ws[r]]
            if all(v is None or str(v).strip() == '' for v in vals):
                continue
            name_val = vals[1] if len(vals) > 1 else None
            if not name_val:
                continue
            
            print(f'--- {str(name_val).strip()} ---')
            for i, val in enumerate(vals):
                if val is not None and str(val).strip() and i < len(headers):
                    print(f'{headers[i]}：{val}')
            print()
