import openpyxl, os, json

dest = r'd:\文旅局数据\文旅局'

def extract_all_data(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.max_row is None or ws.max_row < 8:
            continue
        header_row = None
        for r in range(1, min(15, ws.max_row + 1)):
            row_vals = [str(c.value)[:20] if c.value else '' for c in ws[r]]
            if '序号' in row_vals:
                header_row = r
                break
        if header_row is None:
            continue
        data_start = header_row + 1
        first_data_vals = [str(c.value)[:20] if c.value else '' for c in ws[data_start]]
        if any('样例' in v for v in first_data_vals):
            data_start = header_row + 2
        headers = [c.value for c in ws[header_row]]
        rows = []
        for r in range(data_start, (ws.max_row or 50) + 1):
            vals = [c.value for c in ws[r]]
            if all(v is None or str(v).strip() == '' for v in vals):
                continue
            rows.append(vals)
        if rows:
            print(f'  [{sname}]')
            for i, row in enumerate(rows):
                d = {str(headers[j])[:18]: str(v)[:50] for j, v in enumerate(row) if v is not None and str(v).strip() and j < len(headers)}
                if d:
                    print(f'    {json.dumps(d, ensure_ascii=False)}')

# 嘉乐迪子文件夹
fp = os.path.join(dest, '嘉乐迪', '咸丰县智慧旅游应用数据收集清单-文旅局上报数据模板-2(1).xlsx')
if os.path.exists(fp):
    print(f'\n=== 嘉乐迪 ===')
    extract_all_data(fp)

# 火塘民谣三人行电竞
fp = os.path.join(dest, '火塘民谣 三人行电竞', '咸丰县智慧旅游应用数据收集清单-文旅局上报数据-火塘民谣三人行电竞.xlsx')
if os.path.exists(fp):
    print(f'\n=== 火塘民谣三人行电竞 ===')
    extract_all_data(fp)

# 民宿大文件 (只读数据不读图片)
for fname in ['咸丰县智慧旅游应用数据收集清单-文旅局上报数据模板-民宿(1).xlsx',
              '咸丰县智慧旅游应用数据收集清单-文旅局上报数据模板(8).xlsx',
              '咸丰县智慧旅游应用数据收集清单-文旅局上报数据模板(13).xlsx']:
    fp = os.path.join(dest, fname)
    if os.path.exists(fp):
        print(f'\n=== {fname} ===')
        extract_all_data(fp)
