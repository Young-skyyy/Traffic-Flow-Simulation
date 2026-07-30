import openpyxl, os

dest = r'd:\文旅局数据\文旅局'
fp = os.path.join(dest, '咸丰县智慧旅游应用数据收集清单-文旅局上报数据模板(13).xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)

output_lines = []

def parse_sheet(ws):
    # 找表头行
    header_row = None
    headers = []
    for r in range(1, min(15, ws.max_row + 1)):
        vals = [str(c.value)[:40] if c.value else '' for c in ws[r]]
        if '序号' in vals:
            header_row = r
            headers = [str(c.value) if c.value else '' for c in ws[r]]
            break
    if header_row is None:
        return

    data_start = header_row + 1
    first_vals = [str(c.value)[:20] if c.value else '' for c in ws[data_start]]
    if any('样例' in v for v in first_vals):
        data_start = header_row + 2

    record_count = 0
    for r in range(data_start, (ws.max_row or 50) + 1):
        vals = [c.value for c in ws[r]]
        if all(v is None or str(v).strip() == '' for v in vals):
            continue
        # 跳过分类标签行
        name_val = vals[1] if len(vals) > 1 else None
        if name_val and str(name_val).strip() in ['星级酒店/宾馆', '等级民宿', '红色旅游', '研学基地']:
            continue
        # 跳过提示行
        if name_val and str(name_val).strip() in ['酒店/宾馆/民宿名称', '儿童娱乐区数量']:
            continue
        if not name_val:
            continue

        record_count += 1
        output_lines.append(f"  --- 记录 {record_count} ---")
        for i, val in enumerate(vals):
            if val is not None and str(val).strip() and i < len(headers):
                h = headers[i]
                v = str(val).strip()
                # 只保留必要字段，去掉格式说明行
                if h and v and v not in ['儿童娱乐区数量', '车位数', '规模等级', '酒店/宾馆/民宿名称', '美食街(名称/图片)']:
                    output_lines.append(f"  {h}：{v}")
        output_lines.append("")

for sname in wb.sheetnames:
    ws = wb[sname]
    if ws.max_row is None:
        continue
    output_lines.append(f"{'='*60}")
    output_lines.append(f"【{sname}】")
    output_lines.append(f"{'='*60}")
    parse_sheet(ws)

out_path = r'c:\Users\mnq\Documents\trae_projects\文旅局录入数据.txt'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(output_lines))
print(f'已生成 {out_path}')
print(f'共 {len(output_lines)} 行')
