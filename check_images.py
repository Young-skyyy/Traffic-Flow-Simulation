import openpyxl, os, json

dest = r'd:\文旅局数据\文旅局'

# 用最全的模板(13).xlsx
fp = os.path.join(dest, '咸丰县智慧旅游应用数据收集清单-文旅局上报数据模板(13).xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)

# 列出所有实际存在的图片文件
img_files = []
for root, dirs, files in os.walk(dest):
    for f in files:
        if f.lower().endswith(('.jpg', '.jpeg', '.png', '.mp4', '.docx', '.doc', '.zip')):
            img_files.append(f.lower())
print("=== 已有素材文件 ===")
for f in sorted(set(img_files)):
    print(f"  {f}")

for sname in wb.sheetnames:
    ws = wb[sname]
    if ws.max_row is None:
        continue
    
    # 找表头行
    header_row = None
    headers = []
    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = [str(c.value)[:30] if c.value else '' for c in ws[r]]
        if '序号' in row_vals:
            header_row = r
            headers = [str(c.value)[:20] if c.value else '' for c in ws[r]]
            break
    if header_row is None:
        continue
    
    # 找图片相关列
    img_cols = [i for i, h in enumerate(headers) if any(kw in (h or '') for kw in ['图片', '视频', '文案', '介绍'])]
    
    data_start = header_row + 1
    first_vals = [str(c.value)[:20] if c.value else '' for c in ws[data_start]]
    if any('样例' in v for v in first_vals):
        data_start = header_row + 2

    print(f"\n=== [{sname}] ===")
    no_img = []
    has_img = []
    
    for r in range(data_start, (ws.max_row or 50) + 1):
        vals = [c.value for c in ws[r]]
        if all(v is None or str(v).strip() == '' for v in vals):
            continue
        
        name_val = vals[1] if len(vals) > 1 else None  # 通常第2列是名称
        if name_val is None or str(name_val).strip() == '':
            continue
        name = str(name_val).strip()
        
        # 检查图片相关字段是否有内容
        img_status = {}
        for ci in img_cols:
            if ci < len(vals) and vals[ci] is not None and str(vals[ci]).strip():
                h = headers[ci] if ci < len(headers) else f'Col{ci}'
                img_status[h] = str(vals[ci])[:40]
        
        if img_status:
            has_img.append((name, img_status))
        else:
            no_img.append(name)
    
    if has_img:
        print("  有素材:")
        for name, imgs in has_img:
            print(f"    [{name}] {json.dumps(imgs, ensure_ascii=False)}")
    if no_img:
        print("  缺素材:")
        for name in no_img:
            print(f"    [{name}] 无图片/视频/文案")
